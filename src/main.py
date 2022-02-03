import discord
import os
import openpyxl 
import random
path = "Data.xlsx"
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active 
client = discord.Client()
TOKEN = os.getenv('TOKEN')
@client.event
async def on_ready():
    print('We have logged in as {0.user}'.format(client))
@client.event
async def on_message(message):
    if message.author == client.user:
        return
    if message.content.startswith(''):
        row_x=random.randint(1, 480)
        Poetry = sheet_obj.cell(row = row_x, column = 2) 
        cell_obj2 = sheet_obj.cell(row = row_x, column = 3) 
        Poetry=Poetry.value
        Poetry = Poetry.replace("_x000D_","")
        await message.channel.send(Poetry)
        await message.channel.send(cell_obj2.value)
client.run(TOKEN)
